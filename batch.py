# -*- coding: utf-8 -*-
# @Time    : 2023/1/19  19:29
# @Author  : Leslie
# @File    : batch.py
# @Software: PyCharm
# @Describe:
# -*- encoding:utf-8 -*-
import configparser
import os
from collections import Counter
from optparse import OptionParser
from queue import Queue
from threading import Thread
from urllib.parse import urlparse
from uuid import uuid4
import openpyxl
from IPy import IP
from colorama import Fore
from termcolor import cprint
from tqdm import tqdm

from utils.saveToExcel import saveToExcel
from utils.tools import *

def beian():
    cprint('-' * 50 + 'Load beian ...' + '-' * 50, 'green')
    from spider.subdomain.beian.beian import run_beian
    beianNewDomains, companyName = run_beian(domain)
    for _ in beianNewDomains:
        newDomains.append(_[2])

    # 保存到excel
    beianNewDomainsSheet = saveToExcel(excelSavePath, excel, '备案反查顶级域名')
    beianNewDomainsSheet.saveBeianNewDomains(beianNewDomains)
    return companyName

def Aiqicha(companyName):
    cprint('-' * 50 + 'Load Aiqicha ...' + '-' * 50, 'green')
    if not companyName:
        return

    cprint("查询【{}】公司架构".format(companyName), 'red')

    from spider.subdomain.Aiqicha.Aiqicha import run_Aiqicha
    selfIcpinfo_infos, invest_infos, holds_infos, branch_infos = run_Aiqicha(companyName)

    # 保存到excel
    aiqichaSheet = saveToExcel(excelSavePath, excel, '爱企查')
    aiqichaSheet.saveAiqicha(selfIcpinfo_infos, invest_infos, holds_infos, branch_infos)


def othersApiSubdomain():
    cprint('-' * 50 + 'Load VirusTotal threatcrowd url.fht.im ...' + '-' * 50, 'green')
    from spider.subdomain.otherApi.otherApi import otherApi
    othersApiTotalSubdomains = otherApi(domain,requests_proxies)          # 列表，存放子域名
    return othersApiTotalSubdomains


# 调用kSubdomain脚本
def callKsubdomain():
    cprint('-' * 50 + 'Load ksubdomain ...' + '-' * 50, 'green')
    from spider.subdomain.ksubdomain.ksubdomain import ksubdomain
    ksubdomains = ksubdomain(domain)
    return ksubdomains

def githubApiSubdomain():
    # 获取github敏感信息
    def get_GitSensitiveInfo(github_txt, raw_url_emails):
        conf = configparser.ConfigParser()
        conf.read(getRootPath()+"conf/config.ini",encoding='utf-8')
        secs = conf.sections()
        github_keywords = eval(conf.get('github keywords', 'github_keywords'))

        line_urls = {}
        gitSensitiveInfo = []

        with open(github_txt, 'rt', encoding="utf-8", errors='ignore') as f:
            content = f.readlines()
            for line, each in enumerate(content):
                if '[------------------] ' in each:
                    line_urls[str(line + 1)] = each.split('[------------------] ')[1]

        lines = list(line_urls.keys())

        # print(line_urls)
        def get_githubAddr(line):
            for i, num in enumerate(lines):
                # print(line)
                if i < len(lines) - 1:
                    # print(line, int(num), int(lines[i + 1]))
                    if int(num) <= line <= int(lines[i + 1]):
                        return int(num)
                elif line > int(lines[-1]):
                    return int(lines[-1])

        for keyword in github_keywords:
            for line, each in enumerate(content):
                if line < len(content) - 1:
                    if keyword in each:
                        # print(line)
                        githubAddr = get_githubAddr(line)
                        # print(githubAddr)
                        if githubAddr:
                            raw_url = content[int(githubAddr) - 1].replace('[------------------]', '').strip()
                            try:
                                emails = str(raw_url_emails[raw_url])
                                print('github address: [line:{}] {}'.format(githubAddr, raw_url))
                                print('[emails] : {}'.format(emails))
                                print('[{}] [line:{}] {}'.format(keyword, line, content[line - 1].strip()))
                                print('[{}] [line:{}] {}'.format(keyword, line + 1, content[line].strip()))
                                print('[{}] [line:{}] {}'.format(keyword, line + 2, content[line + 1].strip()))
                                gitSensitiveInfo.append(['gitAddress', githubAddr, raw_url, emails])
                                gitSensitiveInfo.append([keyword, line, content[line - 1].strip(), emails])
                                gitSensitiveInfo.append([keyword, line + 1, content[line].strip(), emails])
                                gitSensitiveInfo.append([keyword, line + 2, content[line + 1].strip(), emails])
                                gitSensitiveInfo.append(['-' * 50, '-' * 50, '-' * 50, '-' * 50])
                            except Exception as e:
                                pass

        return gitSensitiveInfo
    cprint('-' * 50 + 'Load Github Api Subdomain ...' + '-' * 50, 'green')
    from spider.subdomain.githubSubdomain.githubSubdomain import githubSubdomain
    githubApiSubdomains, raw_url_emails = githubSubdomain(domain, save_fold_path,requests_proxies)          # 列表，存放子域名

    # 保存到excel
    githubSheet = saveToExcel(excelSavePath, excel, 'Github敏感信息')
    github_txt = r'{}/{}_github.txt'.format(save_fold_path, domain)
    if os.path.exists(github_txt):
        gitSensitiveInfo = get_GitSensitiveInfo(github_txt, raw_url_emails)
        githubSheet.saveGithub(gitSensitiveInfo)

    return githubApiSubdomains

# 调用爬虫
def SpiderSubdomain():
    cprint('-' * 50 + 'Load Spider ...' + '-' * 50, 'green')  # 启动百度爬虫
    spiderSheet = saveToExcel(excelSavePath, excel, '爬虫')

    # 百度爬虫
    def BaiduSubdomain():
        cprint('Load BaiduSpider ...', 'green')  # 启动百度爬虫
        from spider.subdomain.webSpider.baiduSpider import BaiduSpider
        bdSubdomains, links = BaiduSpider().run_subdomain(domain)
        # 保存到excel
        spiderSheet.saveSpider('百度', links)
        return bdSubdomains

    # 必应爬虫
    def BingSubdomain():
        cprint('Load BingSpider ...', 'green')  # 启动必应爬虫
        from spider.subdomain.webSpider.bingSpider import BingSpider
        bingSubdomains, links = BingSpider().run_subdomain(domain)
        # 保存到excel
        spiderSheet.saveSpider('必应', links)
        return bingSubdomains

    bdSubdomains = BaiduSubdomain()
    bingSubdomains = BingSubdomain()
    spiderSubdomains = list(set(bdSubdomains + bingSubdomains))

    return spiderSubdomains

def crawlCerts(subdomains):
    cprint('-' * 50 + 'Load crawlCerts ...' + '-' * 50, 'green')  # 启动证书爬虫
    from spider.subdomain.crawlCerts.crawlCerts import crawlCerts
    certsSubdomains, trustedDomainDict, _newDomains = crawlCerts(domain, subdomains).run()

    newDomains.extend(_newDomains)
    # 保存到excel
    certSheet = saveToExcel(excelSavePath, excel, '证书')
    certSheet.saveCert(trustedDomainDict)

    return certsSubdomains

# 调用友链爬虫
def FriendChinsSubdomain(temp_subdomains):
    cprint('-' * 50 + 'Load FriendChins ...' + '-' * 50, 'green')  # 启动友链爬虫
    from spider.subdomain.friendChins.friendChins import FriendChins
    fcSubdomains = FriendChins(domain, temp_subdomains).run()
    return fcSubdomains
# 整合子域名，对所有子域名解析A记录
def checkCDN_queryA_subdomains(Subdomains_ips, subdomains):
    cprint('-' * 50 + 'check subdomains CDN and query ip ...' + '-' * 50, 'green')  # 整合所有子域名
    tmp_subdomains = []
    for subdomain in subdomains:
        if '.{}'.format(domain) in subdomain:
            tmp_subdomains.append(subdomain)
    subdomains = list(set(tmp_subdomains))


    print('Check CDN [{}] subdomains'.format(len(subdomains)))
    from spider.subdomain.CDN.checkCDN import checkCDN
    notCDNSubdomains, CDNSubdomainsDict = checkCDN(subdomains)

    print('Query the A record of [{}] subdomains'.format(len(subdomains)))
    from spider.subdomain.queryA.queryA import run_queryA
    Subdomains_ips = run_queryA(Subdomains_ips, subdomains)

    # 保存到excel
    queryASheet = saveToExcel(excelSavePath, excel, '子域名A记录')
    queryASheet.saveQueryA(Subdomains_ips, CDNSubdomainsDict)

    return Subdomains_ips, CDNSubdomainsDict, notCDNSubdomains
# host碰撞
def hostCollide(Subdomains_ips):
    cprint('-' * 50 + 'run_hostCollide ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from spider.subdomain.hostCollide import hostCollide
    hostCollideResult, censysIPS = hostCollide.run_hostCollide(domain, Subdomains_ips)

    # 保存到excel
    queryASheet = saveToExcel(excelSavePath, excel, 'HOST碰撞')
    queryASheet.saveHostCollide(hostCollideResult)

    return censysIPS

# 获取所有子域名的参数链接和后台链接（存活）
def run_ParamLinks():
    cprint('-' * 50 + 'run_ParamLinks ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from spider.subdomain.paramSpider.paramSpider import getParamLinks
    paramLinks, htLinks = getParamLinks(domain)

    # 保存到excel
    paramHtLinksSheet = saveToExcel(excelSavePath, excel, '动态链接和后台地址')
    paramHtLinksSheet.saveparamHtLinks(paramLinks, htLinks)

    # 如果动态链接的个数大于1000，
    if len(paramLinks) > 1000:
        paramLinks = []

    return paramLinks

# 整理IP，获取C段IP
def get_CIP(Subdomains_ips, CDNSubdomainsDict, censysIPS):
    cprint('-' * 50 + 'get_CIP ...' + '-' * 50, 'green')  # 整理C段IP
    # 过滤内网IP
    def is_internal_ip(ip_subnet):
        ip_subnet_list = ip_subnet.split('.')
        if ip_subnet_list[0] == '10' or ip_subnet_list[0] == '127':
            return True
        elif ip_subnet_list[0] == '172' and 15 < int(ip_subnet_list[1]) < 32:
            return True
        elif ip_subnet_list[0] == '192' and ip_subnet_list[1] == '168':
            return True
        else:
            return False

    ips = []
    CIP_List = []
    CIP_List_all = []

    for subdomain in Subdomains_ips:
        if CDNSubdomainsDict[subdomain] == 'NOT':           # 如果该子域名没有CDN，则开始统计解析出来的IP
            ip_List = Subdomains_ips[subdomain]
            for ip in ip_List:
                if not is_internal_ip(ip):
                    ips.append(ip)

    ips.extend(censysIPS)

    for ip in list(set(ips)):
        c_subnet = str(IP(ip).make_net('255.255.255.0')).rsplit('.', 1)[0] + '.0'
        CIP_List_all.append(c_subnet)

    global ip_count
    ip_count = Counter(CIP_List_all)
    cprint(ip_count, 'red')
    import configparser
    cf = configparser.ConfigParser()
    cf.read(getRootPath()+"conf/config.ini",encoding='utf-8')
    c_nums = cf.get('C nums', 'c_nums')

    for ip in ip_count:
        if ip_count[ip] > int(c_nums):
            CIP_List.append(ip)

    return CIP_List
    # return list(set(CIP_List))

def run_subdomain():
    #备案查询
    companyName = beian()
    print(companyName)
    #爱企查
    Aiqicha(companyName)
    Subdomains_ips = {}
    isPanAnalysis = checkPanAnalysis(domain)

    if not isPanAnalysis and ksubdomain:
        # 0. 调用kSubdomain脚本
        # ksubdomains = callKsubdomain()
        ksubdomains = []
    else:
        ksubdomains = []

    print('[total: {}] ksubdomain: {}'.format(len(ksubdomains), ksubdomains))
    subdomains = printGetNewSubdomains([], ksubdomains)
    print('len [{}]'.format(len(subdomains)))

    # 1. theHarvester
    theHarvesterSubdomains = [] #theHarvester()
    print('[total: {}] theHarvester: {}'.format(len(theHarvesterSubdomains), theHarvesterSubdomains))
    subdomains = printGetNewSubdomains(subdomains, theHarvesterSubdomains)
    print('len [{}]'.format(len(subdomains)))
    #
    # # 2. virustotal|ce.baidu.com|www.threatcrowd.org|url.fht.im|qianxun
    othersApiTotalSubdomains = othersApiSubdomain()
    print('[total: {}] webAPI: {}'.format(len(othersApiTotalSubdomains), othersApiTotalSubdomains))
    subdomains = printGetNewSubdomains(subdomains, othersApiTotalSubdomains)
    print('len [{}]'.format(len(subdomains)))

    # 3. github
    githubApiSubdomains = githubApiSubdomain()
    print('[total: {}] Github: {}'.format(len(githubApiSubdomains), githubApiSubdomains))
    subdomains = printGetNewSubdomains(subdomains, githubApiSubdomains)
    #
    # 4. 爬虫(百度｜必应)
    spiderSubdomains = SpiderSubdomain()
    print('[total: {}] Spider: {}'.format(len(spiderSubdomains), spiderSubdomains))
    subdomains = printGetNewSubdomains(subdomains, spiderSubdomains)
    # 防止程序奔溃后重新跑耗费大量时间，所以在当前目录创建文本保存子域名
    with open('{}.txt'.format(domain), 'at') as f:
        for subdomain in subdomains:
            f.writelines('{}\n'.format(subdomain))

    #5.爬证书
    certsSubdomains = crawlCerts(subdomains)
    print('[total: {}] Certs: {}'.format(len(certsSubdomains), certsSubdomains))
    subdomains = printGetNewSubdomains(subdomains, certsSubdomains)

    # 6. 爬友链
    fcSubdomains = FriendChinsSubdomain(subdomains)
    print('[total: {}] Friends: {}'.format(len(fcSubdomains), fcSubdomains))
    subdomains = printGetNewSubdomains(subdomains, fcSubdomains)

    # 防止程序奔溃后重新跑耗费大量时间，所以在当前目录创建文本保存子域名
    with open('{}.txt'.format(domain), 'at') as f:
        for subdomain in subdomains:
            f.writelines('{}\n'.format(subdomain))


    # 7. 整合子域名，对所有子域名判断是否是CDN,解析A记录，并将所有子域名结果保存到excel里
    Subdomains_ips, CDNSubdomainsDict, notCDNSubdomains = checkCDN_queryA_subdomains(Subdomains_ips, subdomains)

    # host碰撞,censysIPS是censys api得到的解析的IP
    censysIPS = hostCollide(Subdomains_ips)

    # 8. 获取所有子域名的参数链接（存活）
    param_Links = run_ParamLinks()

    # 获取C段的IP
    CIP_List = get_CIP(Subdomains_ips, CDNSubdomainsDict, censysIPS)
    print('C段的IP:{}'.format(CIP_List))

    # 8. 跑C段
    run_cSubnet(CIP_List, Subdomains_ips, notCDNSubdomains, param_Links)

def run_webSpace(domain, SubdomainAndNotCDNIPs, CIP_List, fofaTitle):
    cprint('-' * 50 + 'run_webSpace ...' + '-' * 50, 'green')  # 启动网络空间引擎
    from spider.subdomain.webSpace import fofaApi, shodanApi, quakeApi, qianxinApi
    webSpaceSheet = saveToExcel(excelSavePath, excel, '网络空间搜索引擎')
    serviceSheet = saveToExcel(excelSavePath, excel, '服务')
    webSpace_web_host_port = []         # 存放开放web服务
    webSpace_service_host_port = []     # 存放除了Web的其他服务

    # fofa搜索引擎信息收集
    def run_fofa():
        # 查询域名
        if domain:
            query_str = 'domain="{}"'.format(domain)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_domain(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str) # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip="{}/24"'.format(c_subnet)
                fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_ip(query_str)
                if fofa_Results:
                    webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str)
                    webSpace_web_host_port.extend(fofa_web_host_port)
                    webSpace_service_host_port.extend(fofa_service_host_port)

        if fofaTitle:
            query_str = 'title="{}" && country="CN"'.format(fofaTitle)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_domain(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str) # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)



    # shodan搜索引擎信息收集
    def run_shodan():
        # 查询域名
        if domain:
            query_str = 'hostname:"{}"'.format(domain)
            shodan_Results, shodan_web_host_port, shodan_service_host_port = shodanApi.query_domain(query_str)
            if shodan_Results:
                webSpaceSheet.saveWebSpace('shodan', shodan_Results, query_str)
                webSpace_web_host_port.extend(shodan_web_host_port)
                webSpace_service_host_port.extend(shodan_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'net:"{}/24"'.format(c_subnet)
                shodan_Results, shodan_web_host_port, shodan_service_host_port = shodanApi.query_ip(query_str)
                if shodan_Results:
                    webSpaceSheet.saveWebSpace('shodan', shodan_Results, query_str)
                    webSpace_web_host_port.extend(shodan_web_host_port)
                    webSpace_service_host_port.extend(shodan_service_host_port)

    # quake搜索引擎信息收集
    def run_quake():
        # 查询域名
        if domain:
            query_str = 'domain:"{}" AND country:"China"'.format(domain)
            quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_domain(query_str)
            if quake_Results:
                webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                webSpace_web_host_port.extend(quake_web_host_port)
                webSpace_service_host_port.extend(quake_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip:"{}/24"'.format(c_subnet)
                quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_ip(query_str)
                if quake_Results:
                    webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                    webSpace_web_host_port.extend(quake_web_host_port)
                    webSpace_service_host_port.extend(quake_service_host_port)

        if fofaTitle:
            query_str = 'title:"{}" AND country:"China"'.format(fofaTitle)
            quake_Results, quake_web_host_port, quake_service_host_port = quakeApi.query_ip(query_str)
            if quake_Results:
                webSpaceSheet.saveWebSpace('quake', quake_Results, query_str)
                webSpace_web_host_port.extend(quake_web_host_port)
                webSpace_service_host_port.extend(quake_service_host_port)

    # qianxin搜索引擎信息收集
    def run_qianxin():
        # 查询域名
        if domain:
            query_str = '(domain.suffix="{}")&&(country=="中国")'.format(domain)
            qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_domain(query_str)
            if qianxin_Results:
                webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                webSpace_web_host_port.extend(qianxin_web_host_port)
                webSpace_service_host_port.extend(qianxin_service_host_port)

        # 查询C段IP
        if CIP_List:
            for c_subnet in CIP_List:
                query_str = 'ip="{}/24"'.format(c_subnet)
                qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_ip(query_str)
                if qianxin_Results:
                    webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                    webSpace_web_host_port.extend(qianxin_web_host_port)
                    webSpace_service_host_port.extend(qianxin_service_host_port)

        if fofaTitle:
            query_str = '(title="{}")&&(country=="中国")'.format(fofaTitle)
            qianxin_Results, qianxin_web_host_port, qianxin_service_host_port = qianxinApi.query_ip(query_str)
            if qianxin_Results:
                webSpaceSheet.saveWebSpace('qianxin', qianxin_Results, query_str)
                webSpace_web_host_port.extend(qianxin_web_host_port)
                webSpace_service_host_port.extend(qianxin_service_host_port)


    # 对子域名和非CDN的IP进行fofa查询
    def run_fofaOne(subdomainAndIP_Q):
        while not subdomainAndIP_Q.empty():
            subdomainOrIp = subdomainAndIP_Q.get()
            if isIP(subdomainOrIp):
                query_str = 'ip="{}"'.format(subdomainOrIp)
            else:
                query_str = 'domain="{}"'.format(subdomainOrIp)
            fofa_Results, fofa_web_host_port, fofa_service_host_port = fofaApi.query_ip(query_str)
            if fofa_Results:
                webSpaceSheet.saveWebSpace('fofa', fofa_Results, query_str)  # 将网络空间搜索引擎的结果保存到webSpace项里
                # save_webSpace(fofa_Results, 'fofa', query_str)
                webSpace_web_host_port.extend(fofa_web_host_port)
                webSpace_service_host_port.extend(fofa_service_host_port)

    run_fofa()
    run_shodan()
    run_quake()
    run_qianxin()


    # fofa跑所有子域名解析出来的IP
    if SubdomainAndNotCDNIPs:
        subdomainAndIP_Q = Queue(-1)
        for subdomainAndIP in SubdomainAndNotCDNIPs:
            subdomainAndIP_Q.put(subdomainAndIP)
        threads = []
        for t_id in range(5):
            t = Thread(target=run_fofaOne, args=(subdomainAndIP_Q, ))
            threads.append(t)
            t.start()
        for t in threads:
            t.join()

    serviceResult = []
    for _ in webSpace_service_host_port:            # 去重
        if _ not in serviceResult:
            serviceResult.append(_)

    webSpace_service_host_port = serviceResult
    # 将非Web服务的结果保存到service项里
    serviceSheet.saveService(webSpace_service_host_port)


    return webSpace_web_host_port, webSpace_service_host_port


# 整理fofaTitle结果的域名和IP
def collation_fofaDomainIP(webSpace_web_host_port):
    ips = []
    fofaTitle_IPs = []
    fofaTitle_newDomains = []

    for _ in webSpace_web_host_port:
        a = urlparse(_)
        if a.scheme:
            newdomain_ip = a.netloc.split(':')[0]
        else:
            newdomain_ip = a.path.split(':')[0]
        if isIP(newdomain_ip):
            ips.append(newdomain_ip)
        else:
            fofaTitle_newDomains.append(newdomain_ip)

    for ip in list(set(ips)):
        ip_C = str(IP(ip).make_net('255.255.255.0')).rsplit('.', 1)[0] + '.0'
        fofaTitle_IPs.append(ip_C)

    global ip_count
    ip_count = Counter(fofaTitle_IPs)
    newDomains.extend(fofaTitle_newDomains)

# 对IP进行归属地查询
def get_ip_address(web_ip_list):
    cprint('-' * 50 + 'get ip address ...' + '-' * 50, 'green')  # 对IP进行归属地查询
    from spider.subdomain.getIpAddress import getIpAddress
    ip_address_dict = getIpAddress.run_getIpAddress(web_ip_list)            # 字典，key为IP，value为归属地
    return ip_address_dict

# 筛选存活并获取标题
def run_getWebTitle(web_host_port, ip_address_dict):
    tqdm.write(Fore.BLACK + '-' * 50 + 'run_getWebTitle ...' + '-' * 50)  # 筛选存活并获取标题
    from spider.subdomain.webInfo import getWebTitle
    if isIntranet == 1:
        threadNum = 10        # 如果是扫内网，则线程为5
    else:
        threadNum = 300      # 扫外网则线程为300

    # title不需要使用快代理
    requests_proxies = None
    web_Titles = getWebTitle.run_getWebTitle(web_host_port, ip_address_dict, requests_proxies, threadNum)
    # print(web_Titles)
    alive_Web = []  # 存活的web服务
    for each in web_Titles:
        if each[1] != 65535:
            alive_Web.append(each[0])

    # 写入表格里
    webTitileSheet = saveToExcel(excelSavePath, excel, '存活网站标题')  # 创建一个ip反查域名页
    webTitileSheet.saveWebTitle(web_Titles)

    return web_Titles, alive_Web

# ip反查域名，并将域名结果保存到Subdomains_ips列表里，并且存放到ip2domain_dict里
def get_ip2domain():
    cprint('-' * 50 + 'ip to domain ...' + '-' * 50, 'green')  # 对IP进行域名反查
    from spider.subdomain.getIp2Domain import getIp2Domain

    ip2domain_dict, _newDomains = getIp2Domain.run_ip2domain(domain, allTargets_Queue)  # ip2domain_dict字典，key为IP，value为反查的域名

    # 和目标关联的相关域名
    newDomains.extend(_newDomains)

    # 去重
    ip2domainSubdomains = []                        # 反查出来的子域名列表    ['ca.hbu.edu.cn', 'jwjcc.bdu.edu.cn', 'yzuuc.hbu.cn']
    for subdomains in ip2domain_dict.values():
        for subdomain in subdomains:
            if domain:
                if domain in subdomain:
                    ip2domainSubdomains.append(subdomain)
            else:
                ip2domainSubdomains.append(subdomain)
    ip2domainSubdomains = list(set(ip2domainSubdomains))

    ip2domainSheet = saveToExcel(excelSavePath, excel, 'ip反查域名')    # 创建一个ip反查域名页
    ip2domainSheet.saveIp2Domain(ip2domain_dict)

    return ip2domainSubdomains      # 返回ip反查得到的域名列表

# 整理开放web服务的host
def collation_web_host(Subdomains_ips, webSpace_web_host_port, ip2domainSubdomains):
    cprint('-' * 50 + 'collation_web_host ...' + '-' * 50, 'green')  # 启动web服务收集
    web_host_port = []                      # 存放最终的开放web服务的host
    web_host_port_temp = []                 # web_host_port临时列表
    web_ip_list = []                        # 存放开放web服务的IP


    for subdomain in list(set(list(Subdomains_ips.keys()) + ip2domainSubdomains)):
        if ':' in subdomain:                        # ip2domainSubdomains的结果里有些类似于221.192.236.146:999这种结果，所以不加80端口
            web_host_port_temp.append(subdomain)
        else:
            web_host_port_temp.append('{}:80'.format(subdomain))

    web_host_port_temp.extend(webSpace_web_host_port)
    # print('[{}] {}'.format(len(web_host_port), web_host_port))
    web_host_port_temp = list(set(web_host_port_temp))
    # print('[{}] {}'.format(len(web_host_port), web_host_port))

    # 整合url，格式规范。全部是http(s)://www.domain.com:xxxx
    for host_port in web_host_port_temp:
        host_port_urlparse = urlparse(host_port)
        if not host_port_urlparse.scheme:           # 如果没有http（https）， 则加上。如果是443端口，则加https，其他端口加http
            if ':' in host_port:
                try:
                    host, port = host_port.split(':')
                    if isIP(host):
                        web_ip_list.append(host)
                    if port == '443':
                        host_port = 'https://{}'.format(host)
                    elif port == '80':
                        host_port = 'http://{}'.format(host)
                    else:
                        host_port = 'http://{}'.format(host_port)
                except Exception as e:
                    pass
            else:
                host_port = 'http://{}'.format(host_port)
        else:   # 如果有https或者http，则不加
            host_port = host_port
        web_host_port.append(host_port)

    web_host_port = list(set(web_host_port))    # 去重
    web_ip_list = list(set(web_ip_list))

    return web_host_port, web_ip_list

#保存相关信息：新的域名和C段IP信息
def saveRelatedInfo(newDomains, ip_count):
    ip2domainSheet = saveToExcel(excelSavePath, excel, '相关域名和C段')    # 创建一个ip反查域名页
    ip2domainSheet.saveNewDomainAndCSubnet(newDomains, ip_count)

# 获取C段资产
def run_cSubnet(CIP_List, Subdomains_ips, notCDNSubdomains, param_Links):
    print(CIP_List)
    print(Subdomains_ips)
    print(notCDNSubdomains)
    SubdomainAndNotCDNIPs = []  # 子域名和非CDN的IP
    for subdomain in notCDNSubdomains:
        for ip in Subdomains_ips[subdomain]:
            SubdomainAndNotCDNIPs.append(ip)
    SubdomainAndNotCDNIPs = list(set(SubdomainAndNotCDNIPs))
    # 防止IP太多，导致查询次数过多被fofa封
    if len(SubdomainAndNotCDNIPs) > 10:
        SubdomainAndNotCDNIPs = []

    # 8. 调用网络空间引擎,查询根域名和C段IP的资产         webSpace_web_host_port 是Web服务             webSpace_service_host_port  是其他服务
    if domain:            # 跑域名的时候，不跑C段
        webSpace_web_host_port, webSpace_service_host_port = run_webSpace(domain, SubdomainAndNotCDNIPs, [], '')
    else:
        webSpace_web_host_port, webSpace_service_host_port = run_webSpace(domain, [], CIP_List, '')           # 网络空间引擎（fofa、shodan）获取的开放web服务的host（IP/domain）
    # print('webSpace_web_host_port: {}'.format(webSpace_web_host_port))
    # print('webSpace_service_host_port: {}'.format(webSpace_service_host_port))

    for subdomain in Subdomains_ips.keys():
        for ip in Subdomains_ips[subdomain]:
            allTargets_Queue.put(ip)
            allTargets_List.append(ip)

    # ip反查的子域名列表
    ip2domainSubdomains = get_ip2domain()
    print('[total: {}] ip2domainSubdomains: {}'.format(len(ip2domainSubdomains), ip2domainSubdomains))
    print('[ip2domain get new subdomains] [{}]'.format(len(list(set(ip2domainSubdomains)-set(list(Subdomains_ips.keys()))))))

    # 9. 整理开放web服务的host, 存放开放web服务器的ip/domain和port，用来后面的cms识别
    web_host_port, web_ip_list = collation_web_host(Subdomains_ips, webSpace_web_host_port, ip2domainSubdomains)
    print('[total: {}] web_host_port'.format(len(web_host_port)))

    # 10. 对IP进行归属地查询
    ip_address_dict = get_ip_address(web_ip_list)

    # 11. 获取标题, 以及存活的web
    web_Title, alive_Web = run_getWebTitle(web_host_port, ip_address_dict)  # 获取C段资产

    # 不仅仅只信息收集-即跑漏洞
    # if justInfoGather == 0:
    #     webVul_list = detect_webVul(alive_Web)  # 获取C段资产
    #
    #
    #     if domain:
    #         # paramVul_list = detect_paramVul(param_Links)      不跑注入
    #         paramVul_list = []
    #     else:
    #         paramVul_list = []
    #
    #     # 13. 未授权漏洞检测
    #     unauthWeakVul_list = detect_unauthWeakVul(webSpace_service_host_port)       # 获取C段资产
    #     # unauthWeakVul_list = []
    #     # 14. 打印并保存漏洞
    #     Vul_list = webVul_list + unauthWeakVul_list + paramVul_list
    #     printSave_Vul(Vul_list)



    # 15. 保存相关信息：新的域名和C段IP信息
    saveRelatedInfo(newDomains, ip_count)

    cprint(r'新的域名：{}'.format(newDomains), 'green')
    cprint(r'C段IP：{}'.format(CIP_List), 'green')
    cprint(r'资产信息保存路径：{}'.format('{}/{}.xlsx'.format(save_fold_path, excel_name)), 'green')
    cprint(r'Github信息保存路径：{}/{}_github.txt'.format(save_fold_path, domain), 'green')

    if domain:
        ret = ""
        for cip in CIP_List:
            ret += cip
            ret += ","
        cprint(r"请使用-c功能跑C段资产", 'green')

def _init():
    global domain, cSubnet, save_fold_path, excel, excel_name, excelSavePath, proxy, \
        requests_proxies, isIntranet, xlsxFileWB, weak, CIP_List, allTargets_List, \
        allTargets_Queue, masNmapFile, newDomains, ip_count, fofaTitle, ksubdomain, \
        justInfoGather, socksProxysDict, kuaidaili_thread_num
    usage = '\n\t' \
            'python3 %prog -d domain.com\n\t' \
            'python3 %prog -d domain.com --justInfoGather 1\n\t' \
            'python3 %prog -d domain.com --ksubdomain 0\n\t' \
            'python3 %prog -c 192.168.1.0,192.168.2.0,192.168.3.0\n\t' \
            'python3 %prog -f url.txt\n\t' \
            'python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -p 1.1.1.1:1111\n\t' \
            'python3 %prog -n 1 -f url.txt -p 1.1.1.1:1111 --web 1\n\t' \
            'python3 %prog -n 1 -c 192.168.1.0,192.168.2.0 -v 1\n\t' \
            'proxychains4 python3 %prog -n 1 -f /result/2ddcaa3ebbd0/172.18.82.0.xlsx\n\t' \
            'proxychains4 python3 %prog -n 1 -w 1 -f /result/2ddcaa3ebbd0/172.18.82.0.xlsx\n\t' \
            'python3 %prog --mn masNmap.xlsx\n\t' \
            'python3 %prog --mn masNmap.xlsx -w 1\n\t' \
            'python3 %prog --fofaTitle 大学\n\t' \
            'python3 %prog --domainFile domain.txt\n\t'
    parse = OptionParser(usage=usage)
    parse.add_option('-d', '--domain', dest='domain', type='string', help='target domain')
    parse.add_option('-c', '--cSubnet', dest='cSubnet', type='string', help='target cSubnet')
    # parse.add_option('--proxyFlag', dest='proxyFlag', type='int', default=0, help='0:No,1:kuaidaili,2:tencentcs')  # 0不使用代理扫描，1使用快代理扫描，2使用腾讯云函数扫描
    parse.add_option('-n', '--intranet', dest='isIntranet', type='int', default=0, help='Scan intranet value set to 1')        # 扫描内网, 值为1扫内网， 默认为0
    parse.add_option('-p', '--proxy', dest='proxy', type='string', default=None, help='Intranet proxy socks5 socks4')           # 代理，socks5和socks4, 默认为None，可用于外网扫描，也可以用于内网扫描
    parse.add_option('-f', '--file', dest='File', type='string', default=None, help='/result/2ddcaa3ebbd0/172.18.82.0.xlsx')  # 扫描内网的服务漏洞-未授权和弱口令
    parse.add_option('-w', '--weak', dest='weak', type='int', default=None, help='run weak password script')                    # 内网弱口令是否要跑
    parse.add_option('-v', '--vpn', dest='vpn', type='int', default=None, help='Run in the case of vpn')            # 在vpn的情况下跑
    parse.add_option('--web', dest='web', type='int', default=None, help='detect web in Intranet')  # 跑内网的web漏洞
    parse.add_option('--mn', dest='masNmapFile', type='str', default=None, help='run masscan nmap result')          # 跑masscan和nmap的结果
    parse.add_option('--fofaTitle', dest='fofaTitle', type='str', default=None, help='run fofa title')  # 跑fofa的title
    parse.add_option('--domainFile', dest='domainFile', type='str', default=None, help='run domain title')  # 跑多个域名
    parse.add_option('--ksubdomain', dest='ksubdomain', type='int', default=1, help='not run ksubdomain')  # 不使用ksubdomain跑子域名
    parse.add_option('--test', dest='testDemo', type='int', default=0, help='if test=1 then run testDemo')  # 测试某个功能
    parse.add_option('--justInfoGather', dest='justInfoGather', type='int', default=0, help='just infoGather, not detect vul')  # 只信息收集，不跑漏洞
    parse.add_option('--getSocks', dest='getSocks', type='int', default=0, help='get socks')  # 获取socks代理


    options, args = parse.parse_args()
    domain, cSubnet, isIntranet, proxy, File, weak, vpn, masNmapFile, fofaTitle, domainFile, web, ksubdomain, justInfoGather, testDemo, getSocks = options.domain, options.cSubnet, options.isIntranet, options.proxy, options.File, options.weak, options.vpn, options.masNmapFile, options.fofaTitle, options.domainFile, options.web, options.ksubdomain, options.justInfoGather, options.testDemo, options.getSocks
    # 所有目标
    allTargets_List = []
    allTargets_Queue = Queue(-1)

    # C段IP列表
    CIP_List = []

    # C段出现的IP个数
    ip_count = Counter()

    # 和目标资产相关联的新的根域名
    newDomains = []

    # 代理
    socksProxysDict = {"baidu": [], "google": []}
    print(domain, cSubnet, isIntranet, proxy, File)
    #外网代理设置
    if proxy:
        requests_proxies ={"http": "{}".format(proxy), "https": "{}".format(proxy)}
        # print(requests_proxies)

    # 分割C段，获取ip
    if cSubnet:
        CIP_List = cSubnet.split(',')
        for CIP in CIP_List:
            for ip in IP('{}/24'.format(CIP)):
                allTargets_Queue.put(str(ip))
                allTargets_List.append(str(ip))

    # 扫描外网时加载文件扫描
    if File and not isIntranet:
        with open(File, 'rt') as f:
            for each in f.readlines():
                allTargets_Queue.put(each.strip())
                allTargets_List.append(each.strip())

    #创建目录
    #扫描内网漏洞(web或者service)
    if File and isIntranet:
        if not web:     # 扫描内网服务漏洞
            save_fold_path, _excel_name = File.rsplit('/', 1)
            excel_name = _excel_name.rsplit('.', 1)[0] + '_ServiceVul'
            xlsxFileWB = openpyxl.load_workbook(File)  # 打开文件
        else:           # 扫描内网web漏洞
            save_fold_path = os.getcwd() + '/result/' + str(uuid4()).split('-')[-1]  # 保存路径
            os.makedirs(save_fold_path)
            with open(File, 'rt') as f:
                for each in f.readlines():
                    allTargets_Queue.put(each.strip())
                    allTargets_List.append(each.strip())
    #扫描外网或者外网读取file.txt或者读取masNmap.xlsx时
    else:
        try:
            save_fold_path = os.getcwd() + '/result/' + str(uuid4()).split('-')[-1] # 保存路径
            os.makedirs(save_fold_path)
        except Exception:
            pass

    excel = openpyxl.Workbook()
    excel.remove(excel[excel.sheetnames[0]])  # 删除第一个默认的表

    if domain and cSubnet:
        cprint('Error： domain and cSubnet can only pass one', 'red')
        exit(0)
    #跑domain
    elif domain and not cSubnet:
        cprint('-' * 50 + 'Start {} information collection'.format(domain) + '-' * 50, 'green')
        excel_name = domain
        excelSavePath = '{}/{}.xlsx'.format(save_fold_path, excel_name)
        run_subdomain()


if __name__ == '__main__':
    _init()
